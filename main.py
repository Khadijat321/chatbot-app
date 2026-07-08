from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, FileResponse, StreamingResponse, JSONResponse
from pydantic import BaseModel, Field
from groq import Groq
import os
import io
import sys
import uuid
import json
import base64
import requests
import sqlite3
import hashlib
import time
import re
import math
import random
import string
from datetime import datetime
from contextlib import redirect_stdout, redirect_stderr
from typing import Optional, List, Dict, Any, Tuple
import traceback

# Data science & visualization
try:
    import matplotlib  # type: ignore
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt  # type: ignore
    MATPLOTLIB_AVAILABLE = True
except:
    MATPLOTLIB_AVAILABLE = False

try:
    import numpy as np  # type: ignore
    import pandas as pd  # type: ignore
    NUMPY_PANDAS_AVAILABLE = True
except:
    NUMPY_PANDAS_AVAILABLE = False

from io import BytesIO

# Document processing
try:
    # Try the legacy PyPDF2 package first, fall back to the pypdf package if needed
    try:
        import PyPDF2  # type: ignore
    except Exception:
        import pypdf as PyPDF2  # type: ignore
    PYPDF_AVAILABLE = True
except:
    PYPDF_AVAILABLE = False

try:
    from openpyxl import load_workbook  # type: ignore
    OPENPYXL_AVAILABLE = True
except:
    OPENPYXL_AVAILABLE = False

# Text-to-speech
try:
    try:
        from gtts import gTTS  # type: ignore
    except ImportError:
        from gtts import gTTS  # type: ignore
    GTTS_AVAILABLE = True
except:
    GTTS_AVAILABLE = False

# FAISS for RAG
try:
    import faiss  # type: ignore
    FAISS_AVAILABLE = True
except:
    FAISS_AVAILABLE = False

# Sentence transformers for embeddings
try:
    from sentence_transformers import SentenceTransformer
    SENTENCE_TRANSFORMERS_AVAILABLE = True
    embedding_model = SentenceTransformer('all-MiniLM-L6-v2')
except:
    SENTENCE_TRANSFORMERS_AVAILABLE = False
    embedding_model = None

app = FastAPI(title="Omni AI Backend", version="3.0.0")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ========== CONFIGURATION ==========
GROQ_API_KEY = os.getenv("GROQ_API_KEY", "")
HUGGINGFACE_API_KEY = os.getenv("HUGGINGFACE_API_KEY", "")
TAVILY_API_KEY = os.getenv("TAVILY_API_KEY", "")
NEWSDATA_API_KEY = os.getenv("NEWSDATA_API_KEY", "")
OPENWEATHER_API_KEY = os.getenv("OPENWEATHER_API_KEY", "")
SERPAPI_KEY = os.getenv("SERPAPI_KEY", "")

# Store data
conversations: Dict[str, List[Dict]] = {}
generated_files: Dict[str, str] = {}
user_memories: Dict[str, Dict] = {}

# RAG document store with vector embeddings
rag_documents: Dict[str, List[Dict]] = {}
rag_embeddings: Dict[str, Any] = {}

# Initialize SQLite for persistent memory
DB_PATH = "omni_ai_memory.db"

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    
    c.execute('''
        CREATE TABLE IF NOT EXISTS conversations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT,
            role TEXT,
            content TEXT,
            timestamp REAL,
            mode TEXT
        )
    ''')
    
    c.execute('''
        CREATE TABLE IF NOT EXISTS documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT,
            filename TEXT,
            content TEXT,
            chunks TEXT,
            timestamp REAL
        )
    ''')
    
    c.execute('''
        CREATE TABLE IF NOT EXISTS generated_content (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT,
            type TEXT,
            content TEXT,
            url TEXT,
            timestamp REAL
        )
    ''')
    
    c.execute('''
        CREATE TABLE IF NOT EXISTS user_memories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT,
            key TEXT,
            value TEXT,
            timestamp REAL
        )
    ''')
    
    c.execute('''
        CREATE TABLE IF NOT EXISTS academic_cache (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            query TEXT,
            papers TEXT,
            timestamp REAL
        )
    ''')
    
    conn.commit()
    conn.close()

init_db()

# ========== MODELS ==========
class ChatRequest(BaseModel):
    message: str
    session_id: str = "default"
    mode: str = "chat"
    image_data: Optional[str] = None

class AgentRequest(BaseModel):
    message: str
    session_id: str = "default"
    tools: List[str] = Field(default_factory=lambda: ["search", "code", "weather", "news", "academic", "financial"])

class TranslateRequest(BaseModel):
    message: str
    target_language: str = "en"
    session_id: str = "default"

class SentimentRequest(BaseModel):
    message: str
    session_id: str = "default"

class ChartRequest(BaseModel):
    message: str
    session_id: str = "default"

class WeatherRequest(BaseModel):
    message: str
    session_id: str = "default"

class NewsRequest(BaseModel):
    message: str
    session_id: str = "default"

class SearchRequest(BaseModel):
    message: str
    session_id: str = "default"

class AcademicRequest(BaseModel):
    message: str
    session_id: str = "default"
    max_results: int = 5

class FinancialRequest(BaseModel):
    message: str
    session_id: str = "default"
    symbol: Optional[str] = None

class TTSRequest(BaseModel):
    message: str
    lang: str = "en"
    session_id: str = "default"

class ImageRequest(BaseModel):
    message: str
    session_id: str = "default"

class ImageSearchRequest(BaseModel):
    message: str
    session_id: str = "default"
    num_results: int = 5

class WebsiteRequest(BaseModel):
    message: str
    session_id: str = "default"

class AnalyzeRequest(BaseModel):
    message: str
    session_id: str = "default"

class VideoRequest(BaseModel):
    message: str
    session_id: str = "default"

class RAGQueryRequest(BaseModel):
    message: str
    session_id: str = "default"

class CreativeRequest(BaseModel):
    message: str
    session_id: str = "default"
    genre: str = "story"

class StudyGuideRequest(BaseModel):
    message: str
    session_id: str = "default"
    level: str = "intermediate"

class DebateRequest(BaseModel):
    message: str
    session_id: str = "default"
    stance: Optional[str] = None

class ClearRequest(BaseModel):
    session_id: str = "default"

# ========== HELPER FUNCTIONS ==========
def get_client():
    return Groq(api_key=GROQ_API_KEY)

def save_message(session_id: str, role: str, content: str, mode: str = 'chat'):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        "INSERT INTO conversations (session_id, role, content, timestamp, mode) VALUES (?, ?, ?, ?, ?)",
        (session_id, role, content, time.time(), mode)
    )
    conn.commit()
    conn.close()

def get_conversation_history(session_id: str, limit: int = 20):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        "SELECT role, content FROM conversations WHERE session_id = ? ORDER BY timestamp DESC LIMIT ?",
        (session_id, limit)
    )
    rows = c.fetchall()
    conn.close()
    return [{"role": r[0], "content": r[1]} for r in reversed(rows)]

def save_memory(session_id: str, key: str, value: str):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        "INSERT OR REPLACE INTO user_memories (session_id, key, value, timestamp) VALUES (?, ?, ?, ?)",
        (session_id, key, value, time.time())
    )
    conn.commit()
    conn.close()

def get_memories(session_id: str) -> Dict[str, str]:
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT key, value FROM user_memories WHERE session_id = ?", (session_id,))
    rows = c.fetchall()
    conn.close()
    return {r[0]: r[1] for r in rows}

# ========== WEB SEARCH (Enhanced) ==========
def web_search(query: str, num_results: int = 5) -> str:
    """Search the web using Tavily with DuckDuckGo fallback"""
    if TAVILY_API_KEY:
        try:
            url = "https://api.tavily.com/search"
            headers = {"Content-Type": "application/json"}
            payload = {
                "api_key": TAVILY_API_KEY,
                "query": query,
                "search_depth": "basic",
                "max_results": num_results,
                "include_answer": True
            }

            resp = requests.post(url, headers=headers, json=payload, timeout=10)
            data = resp.json()

            answer = data.get("answer", "")
            output = f"\n**Web Search Results for '{query}':**\n"

            if answer:
                output += f"\n**Quick Answer:** {answer}\n"

            for i, result in enumerate(data.get("results", [])[:num_results], 1):
                title = result.get("title", "No title")
                content = result.get("content", "")
                url_link = result.get("url", "")

                output += f"\n{i}. **{title}**\n"
                if content:
                    output += f"   {content[:300]}{'...' if len(content) > 300 else ''}\n"
                if url_link:
                    output += f"   {url_link}\n"

            return output

        except Exception as e:
            pass

    # DuckDuckGo fallback
    try:
        url = f"https://html.duckduckgo.com/html/?q={requests.utils.quote(query)}"
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        resp = requests.get(url, headers=headers, timeout=10)

        titles = re.findall(r'<a[^>]+class="result__a"[^>]*>(.*?)</a>', resp.text)
        snippets = re.findall(r'<a[^>]+class="result__snippet"[^>]*>(.*?)</a>', resp.text)

        output = f"\n**Web Search Results for '{query}':**\n"
        for i in range(min(num_results, len(titles))):
            title = re.sub(r'<[^>]+>', '', titles[i])
            snippet = re.sub(r'<[^>]+>', '', snippets[i]) if i < len(snippets) else ""
            output += f"\n{i+1}. **{title}**\n   {snippet}\n"

        return output if titles else f"\nSearched for '{query}'. No results found.\n"

    except Exception as e:
        return f"\nWeb search unavailable. Add TAVILY_API_KEY for better search.\n"

# ========== ACADEMIC SEARCH ==========
def search_academic(query: str, max_results: int = 5) -> str:
    """Search academic papers using arXiv and Semantic Scholar"""
    results = []
    
    # Try arXiv first
    try:
        arxiv_url = f"http://export.arxiv.org/api/query?search_query=all:{requests.utils.quote(query)}&start=0&max_results={max_results}&sortBy=relevance&sortOrder=descending"
        resp = requests.get(arxiv_url, timeout=15)
        
        entries = re.findall(r'<entry>(.*?)</entry>', resp.text, re.DOTALL)
        for entry in entries[:max_results]:
            title = re.search(r'<title>(.*?)</title>', entry, re.DOTALL)
            summary = re.search(r'<summary>(.*?)</summary>', entry, re.DOTALL)
            authors = re.findall(r'<name>(.*?)</name>', entry)
            pdf_link = re.search(r'<link[^>]*title="pdf"[^>]*href="([^"]*)"', entry)
            
            if title:
                results.append({
                    "source": "arXiv",
                    "title": title.group(1).strip().replace("\n", " "),
                    "abstract": (summary.group(1).strip()[:400] + "...") if summary else "No abstract",
                    "authors": ", ".join(authors[:3]),
                    "url": pdf_link.group(1) if pdf_link else ""
                })
    except Exception as e:
        pass
    
    # Try Semantic Scholar
    try:
        ss_url = f"https://api.semanticscholar.org/graph/v1/paper/search?query={requests.utils.quote(query)}&fields=title,abstract,authors,year,url&limit={max_results}"
        resp = requests.get(ss_url, timeout=10)
        data = resp.json()
        
        for paper in data.get("data", [])[:max_results]:
            authors = [a.get("name", "") for a in paper.get("authors", [])[:3]]
            results.append({
                "source": "Semantic Scholar",
                "title": paper.get("title", "No title"),
                "abstract": (paper.get("abstract", "")[:400] + "...") if paper.get("abstract") else "No abstract",
                "authors": ", ".join(authors),
                "year": paper.get("year", ""),
                "url": paper.get("url", "")
            })
    except Exception as e:
        pass
    
    if not results:
        return f"\nNo academic papers found for '{query}'. Try a more specific query.\n"
    
    output = f"\n**Academic Papers on '{query}':**\n"
    for i, paper in enumerate(results[:max_results], 1):
        output += f"\n{i}. **{paper['title']}** ({paper['source']})\n"
        output += f"   Authors: {paper['authors']}\n"
        if paper.get("year"):
            output += f"   Year: {paper['year']}\n"
        output += f"   {paper['abstract']}\n"
        if paper.get("url"):
            output += f"   [Read more]({paper['url']})\n"
    
    return output

# ========== FINANCIAL DATA ==========
def get_financial_data(symbol: str, data_type: str = "quote") -> str:
    """Get financial data using Yahoo Finance (free)"""
    try:
        if data_type == "quote":
            url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?interval=1d&range=1mo"
            resp = requests.get(url, timeout=10)
            data = resp.json()
            
            result = data.get("chart", {}).get("result", [{}])[0]
            meta = result.get("meta", {})
            
            price = meta.get("regularMarketPrice", "N/A")
            prev_close = meta.get("previousClose", "N/A")
            currency = meta.get("currency", "USD")
            
            change = ""
            if price != 'N/A' and prev_close != 'N/A':
                try:
                    change_pct = ((float(price) - float(prev_close)) / float(prev_close)) * 100
                    change = f" ({change_pct:+.2f}%)"
                except:
                    pass
            
            return f"**{symbol.upper()}**\nCurrent Price: {price} {currency}{change}\nPrevious Close: {prev_close} {currency}"
        
        elif data_type == "history":
            url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?interval=1d&range=6mo"
            resp = requests.get(url, timeout=10)
            data = resp.json()
            
            result = data.get("chart", {}).get("result", [{}])[0]
            timestamps = result.get("timestamp", [])
            prices = result.get("indicators", {}).get("quote", [{}])[0].get("close", [])
            
            if timestamps and prices:
                latest = prices[-1]
                earliest = prices[0]
                try:
                    change = ((latest - earliest) / earliest) * 100
                    return f"**{symbol.upper()} - 6 Month Performance**\nStart: {earliest:.2f} | End: {latest:.2f} | Change: {change:+.2f}%"
                except:
                    pass
            
    except Exception as e:
        return f"Financial data fetch failed for {symbol}: {str(e)}"
    
    return f"No data found for {symbol}"

# ========== WEATHER ==========
def get_weather(city: str) -> str:
    """Get weather data from OpenWeatherMap"""
    if not OPENWEATHER_API_KEY:
        return "Weather API not configured. Add OPENWEATHER_API_KEY."
    try:
        url = f"http://api.openweathermap.org/data/2.5/weather?q={city}&appid={OPENWEATHER_API_KEY}&units=metric"
        resp = requests.get(url, timeout=10)
        data = resp.json()
        if data.get("cod") != 200:
            return f"Weather error: {data.get('message', 'Unknown error')}"

        temp = data["main"]["temp"]
        feels_like = data["main"]["feels_like"]
        humidity = data["main"]["humidity"]
        desc = data["weather"][0]["description"]
        wind = data["wind"]["speed"]
        pressure = data["main"]["pressure"]
        visibility = data.get("visibility", 0) / 1000

        return f"""**Weather in {city.title()}**
🌡️ Temperature: {temp}°C (feels like {feels_like}°C)
💧 Humidity: {humidity}%
☁️ Conditions: {desc.title()}
💨 Wind: {wind} m/s
📊 Pressure: {pressure} hPa
👁️ Visibility: {visibility:.1f} km"""
    except Exception as e:
        return f"Weather fetch failed: {str(e)}"

# ========== NEWS ==========
def get_news(query: str = "technology", num: int = 5) -> str:
    """Get news from NewsData.io with fallback to web search"""
    if NEWSDATA_API_KEY:
        try:
            url = f"https://newsdata.io/api/1/latest?apikey={NEWSDATA_API_KEY}&q={requests.utils.quote(query)}&language=en&size={num}"
            resp = requests.get(url, timeout=10)
            data = resp.json()

            if data.get("status") == "success":
                articles = data.get("results", [])
                output = f'\n**Latest News on "{query}":**\n'

                for i, art in enumerate(articles[:num], 1):
                    title = art.get("title", "No title")
                    desc = art.get("description", "") or art.get("content", "")[:200]
                    link = art.get("link", "")
                    source = art.get("source_id", "Unknown")
                    pub_date = art.get("pubDate", "")

                    output += f"\n{i}. **{title}**\n"
                    if desc:
                        output += f"   {desc}\n"
                    if link:
                        output += f"   {link}\n"
                    output += f"   📰 {source}"
                    if pub_date:
                        output += f" | {pub_date}"
                    output += "\n"

                return output
        except Exception as e:
            pass
    
    # Fallback to web search for news
    return web_search(f"latest news {query}", num_results=num)

# ========== IMAGE GENERATION ==========
def generate_image(prompt: str, width: int = 1024, height: int = 768) -> str:
    """Generate image using Pollinations.ai (free)"""
    clean = requests.utils.quote(prompt[:400])
    return f"https://image.pollinations.ai/prompt/{clean}?width={width}&height={height}&nologo=true&seed={int(time.time())}&enhance=true"

# ========== IMAGE SEARCH ==========
def search_images(query: str, num: int = 5) -> List[str]:
    """Search for images online using DuckDuckGo image search"""
    try:
        url = f"https://duckduckgo.com/?q={requests.utils.quote(query)}&iax=images&ia=images"
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        resp = requests.get(url, headers=headers, timeout=10)
        
        image_urls = re.findall(r'https://[^"\s]+\.(?:jpg|jpeg|png|gif|webp)', resp.text)
        return list(set(image_urls))[:num]
    except:
        return []

# ========== CHART GENERATION ==========
def generate_chart(chart_type: str, data: dict, title: str = "Chart") -> str:
    """Generate matplotlib chart and return base64"""
    try:
        plt.style.use('dark_background')
        fig, ax = plt.subplots(figsize=(10, 6))
        fig.patch.set_facecolor('#0a0a12')
        ax.set_facecolor('#12121c')

        if chart_type == "bar":
            bars = ax.bar(data.get("labels", []), data.get("values", []), color="#6366f1", edgecolor="#818cf8", linewidth=1.5)
            for bar in bars:
                bar.set_alpha(0.85)
        elif chart_type == "line":
            ax.plot(data.get("labels", []), data.get("values", []), marker="o", color="#6366f1", linewidth=2.5, markersize=8)
            ax.fill_between(range(len(data.get("values", []))), data.get("values", []), alpha=0.15, color="#6366f1")
        elif chart_type == "pie":
            colors = ["#6366f1", "#10b981", "#f59e0b", "#ef4444", "#8b5cf6", "#06b6d4", "#ec4899"]
            wedges, texts, autotexts = ax.pie(data.get("values", []), labels=data.get("labels", []), autopct="%1.1f%%",
                                              colors=colors[:len(data.get('values', []))], startangle=90)
            for autotext in autotexts:
                autotext.set_color('white')
                autotext.set_fontweight('bold')
        elif chart_type == "scatter":
            ax.scatter(data.get("x", []), data.get("y", []), color="#6366f1", alpha=0.7, s=80, edgecolors="#818cf8", linewidths=1)
        elif chart_type == "histogram":
            ax.hist(data.get("values", []), bins=data.get("bins", 10), color="#6366f1", alpha=0.7, edgecolor="#818cf8")
        elif chart_type == "area":
            ax.fill_between(data.get("labels", []), data.get("values", []), alpha=0.3, color="#6366f1")
            ax.plot(data.get("labels", []), data.get("values", []), color="#6366f1", linewidth=2)

        ax.set_title(title, fontsize=16, fontweight='bold', color='white', pad=15)
        ax.set_xlabel(data.get('xlabel', ''), color='#94a3b8', fontsize=11)
        ax.set_ylabel(data.get('ylabel', ''), color='#94a3b8', fontsize=11)
        ax.tick_params(colors='#94a3b8', labelsize=10)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_color('#232335')
        ax.spines['bottom'].set_color('#232335')
        ax.grid(True, alpha=0.1, color='white')

        plt.tight_layout()
        buf = BytesIO()
        plt.savefig(buf, format="png", dpi=120, bbox_inches="tight", facecolor="#0a0a12")
        plt.close()
        buf.seek(0)
        img_base64 = base64.b64encode(buf.read()).decode()
        return f"data:image/png;base64,{img_base64}"
    except Exception as e:
        return f"Error generating chart: {str(e)}"

# ========== TEXT TO SPEECH ==========
def text_to_speech(text: str, lang: str = "en") -> str:
    """Convert text to speech, return base64 audio"""
    if not GTTS_AVAILABLE:
        return None
    try:
        tts = gTTS(text=text[:500], lang=lang, slow=False)
        buf = BytesIO()
        tts.write_to_fp(buf)
        buf.seek(0)
        audio_b64 = base64.b64encode(buf.read()).decode()
        return f"data:audio/mp3;base64,{audio_b64}"
    except Exception as e:
        return None

# ========== FILE EXTRACTION ==========
def extract_text_from_file(content: bytes, filename: str) -> str:
    """Extract text from various file types"""
    ext = filename.lower().split(".")[-1] if "." in filename else ""

    if ext in ["txt", "md", "csv", "json", "py", "js", "html", "css"]:
        return content.decode("utf-8", errors="ignore")

    elif ext == "pdf" and PYPDF_AVAILABLE:
        try:
            reader = PyPDF2.PdfReader(BytesIO(content))
            text = ''
            for page in reader.pages:
                text += page.extract_text() or ''
            return text
        except:
            return "Error reading PDF"

    elif ext in ["xlsx", "xls"] and OPENPYXL_AVAILABLE:
        try:
            wb = load_workbook(BytesIO(content))
            text = ''
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                text += f"\n--- Sheet: {sheet} ---\n"
                for row in ws.iter_rows(values_only=True):
                    text += " | ".join(str(c) if c else "" for c in row) + "\n"
            return text
        except:
            return "Error reading Excel file"

    return content.decode("utf-8", errors="ignore")[:10000]

# ========== RAG WITH EMBEDDINGS ==========
def chunk_text(text: str, chunk_size: int = 500, overlap: int = 50) -> List[str]:
    """Split text into overlapping chunks"""
    words = text.split()
    chunks = []
    for i in range(0, len(words), chunk_size - overlap):
        chunk = " ".join(words[i:i + chunk_size])
        if chunk:
            chunks.append(chunk)
    return chunks

def create_rag_index(session_id: str, documents: List[Dict]):
    """Create FAISS index for RAG"""
    if not FAISS_AVAILABLE or not SENTENCE_TRANSFORMERS_AVAILABLE:
        return
    
    all_chunks = []
    chunk_metadata = []
    
    for doc in documents:
        chunks = chunk_text(doc['content'])
        for i, chunk in enumerate(chunks):
            all_chunks.append(chunk)
            chunk_metadata.append({"filename": doc["filename"], "chunk_index": i})
    
    if not all_chunks:
        return
    
    embeddings = embedding_model.encode(all_chunks)
    dimension = embeddings.shape[1]
    
    index = faiss.IndexFlatL2(dimension)
    index.add(np.array(embeddings).astype('float32'))
    
    rag_embeddings[session_id] = {
        "index": index,
        "chunks": all_chunks,
        "metadata": chunk_metadata
    }

def search_rag(session_id: str, query: str, top_k: int = 5) -> List[Tuple[str, str]]:
    """Search RAG index for relevant chunks"""
    if session_id not in rag_embeddings or not SENTENCE_TRANSFORMERS_AVAILABLE:
        return []
    
    query_embedding = embedding_model.encode([query])
    index_data = rag_embeddings[session_id]
    
    distances, indices = index_data['index'].search(
        np.array(query_embedding).astype('float32'), top_k
    )
    
    results = []
    for idx in indices[0]:
        if idx < len(index_data['chunks']):
            chunk = index_data['chunks'][idx]
            meta = index_data['metadata'][idx]
            results.append((meta["filename"], chunk))
    
    return results

# ========== SYSTEM PROMPTS (Enhanced) ==========
SYSTEM_PROMPTS = {
    "chat": "You are Omni AI, a super-intelligent assistant with access to tools, vision, code execution, web search, image generation, data analysis, academic research, financial data, and more. Be helpful, concise, accurate, and engaging. When uncertain, say so. Always cite sources when using web search or academic data.",
    "code": "You are an expert programmer. Write clean, efficient, well-documented code. Always include error handling. Explain complex logic with comments. Follow PEP 8 for Python.",
    "vision": "You are a computer vision expert. Describe images in detail, identify objects, read text, analyze scenes, and answer questions about visual content. Be precise and thorough.",
    "agent": "You are an AI agent that intelligently selects and uses tools. Analyze user requests, determine which tools are needed (search, code, weather, news, academic, financial, image, chart), execute them, and synthesize comprehensive responses.",
    "analyst": "You are a data scientist. Analyze data thoroughly, identify trends, create insights, suggest visualizations, and provide actionable recommendations. Use statistical reasoning.",
    "creative": "You are a creative writing expert. Write engaging, original stories, poems, scripts, and essays. Develop believable characters, immersive settings, and compelling plots. Adapt tone and style to the genre.",
    "debate": "You are a skilled debater and negotiator. Present well-reasoned arguments, consider multiple perspectives, weigh pros and cons, and address counter-arguments. Be persuasive but fair.",
    "emotional": "You are emotionally intelligent. Recognize emotions in text, empathize with users, and provide compassionate, personalized responses. Be supportive and understanding.",
    "study": "You are an expert educator. Create clear study guides, explain complex concepts at the user level, use analogies, and provide practice questions. Adapt to beginner/intermediate/advanced levels.",
    "research": "You are a research assistant. Synthesize information from multiple sources, provide balanced perspectives, cite sources, and identify gaps in knowledge. Be thorough and objective."
}
# ========== CORE ENDPOINTS ==========

@app.get("/")
async def root():
    return {
        "name": "Omni AI Backend v3.0",
        "status": "Active",
        "features": [
            "chat", "vision", "code_execution", "website_builder",
            "data_analysis", "file_upload", "image_generation", "image_search",
            "web_search", "weather", "news", "agent_mode", "academic_search",
            "financial_data", "text_to_speech", "chart_generation", "youtube_creator",
            "translation", "sentiment_analysis", "rag_qa", "creative_writing",
            "study_guides", "debate_negotiation", "emotional_intelligence"
        ],
        "models": ["llama-3.1-8b-instant", "llama-3.2-90b-vision-preview", "whisper-large-v3"],
        "version": "3.0.0"
    }

# ========== 1. ENHANCED CHAT ==========
@app.post("/chat")
async def chat(request: ChatRequest):
    session_id = request.session_id or "default"

    if session_id not in conversations:
        conversations[session_id] = [
            {"role": "system", "content": SYSTEM_PROMPTS.get(request.mode, SYSTEM_PROMPTS["chat"])}
        ]

    if request.image_data:
        return await vision_chat(request)

    conversations[session_id].append({"role": "user", "content": request.message})
    save_message(session_id, "user", request.message, request.mode)

    # Load user memories for personalization
    memories = get_memories(session_id)
    memory_context = ""
    if memories:
        memory_context = "\nUser preferences: " + "; ".join([f"{k}: {v}" for k, v in list(memories.items())[:5]])

    try:
        client = get_client()
        messages = conversations[session_id].copy()
        if memory_context:
            messages[-1]["content"] = memory_context + "\n\n" + messages[-1]["content"]

        response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=messages,
            temperature=0.7,
            max_tokens=4096
        )
        reply = response.choices[0].message.content

        conversations[session_id].append({"role": "assistant", "content": reply})
        save_message(session_id, "assistant", reply, request.mode)

        return {"reply": reply, "mode": request.mode, "session_id": session_id}
    except Exception as e:
        return {"reply": f"Error: {str(e)}", "mode": request.mode}

# ========== 2. VISION AI ==========
@app.post("/vision")
async def vision_chat(request: ChatRequest):
    session_id = request.session_id or "default"

    if not request.image_data:
        return {"reply": "No image provided. Please upload an image."}

    try:
        client = get_client()

        image_content = request.image_data
        if image_content.startswith("data:image"):
            image_content = image_content.split(",")[1]

        messages = [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": request.message or "Describe this image in detail."},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_content}"}}
                ]
            }
        ]

        response = client.chat.completions.create(
            model="llama-3.2-90b-vision-preview",
            messages=messages,
            max_tokens=4096
        )

        reply = response.choices[0].message.content
        save_message(session_id, "user", f"[Image] {request.message}", "vision")
        save_message(session_id, "assistant", reply, "vision")

        return {"reply": reply, "mode": "vision"}
    except Exception as e:
        return {"reply": f"Vision analysis error: {str(e)}"}

# ========== 3. ENHANCED CODE EXECUTION ==========
@app.post("/execute")
async def execute_code(request: ChatRequest):
    try:
        client = get_client()

        code_prompt = f"""Write only executable Python code for this request. No explanations outside code blocks:

Request: {request.message}

Rules:
- Write clean, working Python code
- Use print() to show output
- Handle errors with try/except
- For charts, save to buffer and print base64
- If creating files, use simple names in /tmp/"""

        response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": code_prompt}],
            temperature=0.3
        )

        raw_code = response.choices[0].message.content

        code = raw_code
        if "```python" in raw_code:
            code = raw_code.split("```python")[1].split("```")[0].strip()
        elif "```" in raw_code:
            code = raw_code.split("```")[1].split("```")[0].strip()

        safe_globals = {
            "__builtins__": {
                "print": print, "len": len, "range": range,
                "str": str, "int": int, "float": float, "bool": bool,
                "list": list, "dict": dict, "set": set, "tuple": tuple,
                "zip": zip, "map": map, "filter": filter, "sum": sum,
                "min": min, "max": max, "abs": abs, "round": round,
                "sorted": sorted, "enumerate": enumerate, "reversed": reversed,
                "type": type, "isinstance": isinstance, "hasattr": hasattr,
                "getattr": getattr, "setattr": setattr,
                "Exception": Exception, "ValueError": ValueError,
                "KeyError": KeyError, "IndexError": IndexError,
                "ZeroDivisionError": ZeroDivisionError,
                "open": open, "input": lambda x="": "",
                "json": json, "os": os, "sys": sys,
                "math": __import__("math"),
                "random": __import__("random"),
                "datetime": __import__("datetime"),
                "re": __import__("re"),
                "itertools": __import__("itertools"),
                "collections": __import__("collections"),
                "statistics": __import__("statistics"),
                "hashlib": hashlib,
                "base64": base64,
                "io": io,
                "requests": requests,
                "numpy": np,
                "pandas": pd,
                "plt": plt,
                "matplotlib": matplotlib,
            },
            "np": np, "pd": pd, "plt": plt, "matplotlib": matplotlib,
            "BytesIO": BytesIO, "base64": base64,
        }

        output_buffer = io.StringIO()
        error_buffer = io.StringIO()

        with redirect_stdout(output_buffer), redirect_stderr(error_buffer):
            try:
                exec(code, safe_globals)
            except Exception as e:
                print(f"Runtime Error: {type(e).__name__}: {str(e)}")

        output = output_buffer.getvalue()
        errors = error_buffer.getvalue()

        full_output = output if output.strip() else "Code executed successfully (no output)"
        if errors.strip():
            full_output += f"\n\nStderr:\n{errors}"

        return {
            "reply": f"""**Code Generated & Executed!**

```python
{code}
```

**Output:**
```
{full_output}
```""",
            "code": code,
            "output": full_output
        }
    except Exception as e:
        return {"reply": f"Code execution error: {str(e)}"}

# ========== 4. WEBSITE BUILDER ==========
@app.post("/build-website")
async def build_website(request: WebsiteRequest):
    try:
        client = get_client()

        website_prompt = f"""Create a complete, self-contained HTML website based on this request.

Request: {request.message}

Requirements:
- Create a single HTML file with embedded CSS and JavaScript
- Use modern, clean design with good UX
- Make it fully functional and interactive
- Include all necessary styles inline
- Use placeholder images from picsum.photos or similar if needed
- Return ONLY the complete HTML code, no explanations

Return the complete HTML code:"""

        response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": website_prompt}],
            temperature=0.7,
            max_tokens=4096
        )

        html_code = response.choices[0].message.content

        if "```html" in html_code:
            html_code = html_code.split("```html")[1].split("```")[0].strip()
        elif "```" in html_code:
            html_code = html_code.split("```")[1].split("```")[0].strip()

        file_id = str(uuid.uuid4())[:8]
        filename = f"/tmp/website_{file_id}.html"
        with open(filename, "w", encoding="utf-8") as f:
            f.write(html_code)

        generated_files[request.session_id] = filename
        preview_url = f"/preview/{file_id}"

        return {
            "reply": "Website created successfully! You can download or preview it below.",
            "code": html_code,
            "preview_url": preview_url,
            "file_id": file_id
        }
    except Exception as e:
        return {"reply": f"Website build error: {str(e)}"}

@app.get("/preview/{file_id}")
async def preview_website(file_id: str):
    for sid, path in generated_files.items():
        if file_id in path:
            return HTMLResponse(content=open(path, "r", encoding="utf-8").read())
    return {"error": "Website not found"}

# ========== 5. DATA ANALYSIS ==========
@app.post("/analyze")
async def analyze_data(request: AnalyzeRequest):
    try:
        client = get_client()

        analyze_prompt = f"""Analyze the following data/text and provide insights.

Data: {request.message}

Please provide:
1. Summary of the data
2. Key statistics and metrics
3. Trends or patterns
4. Actionable insights
5. Suggestions for visualization if applicable

Format your response with clear sections."""

        response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": analyze_prompt}],
            temperature=0.5,
            max_tokens=4096
        )

        reply = response.choices[0].message.content

        # Try to generate a chart if data looks like it has numbers
        chart_image = None
        try:
            numbers = re.findall(r'\d+\.?\d*', request.message)
            if len(numbers) >= 3:
                chart_prompt = f"""Based on this data: {request.message}

Create a Python dictionary for a bar chart with this exact format:
{{"labels": ["label1", "label2", ...], "values": [val1, val2, ...], "title": "Chart Title"}}

Return ONLY the dictionary, no other text."""

                chart_response = client.chat.completions.create(
                    model="llama-3.1-8b-instant",
                    messages=[{"role": "user", "content": chart_prompt}],
                    temperature=0.3
                )

                chart_text = chart_response.choices[0].message.content
                dict_match = re.search(r'\{.*\}', chart_text, re.DOTALL)
                if dict_match:
                    chart_data = eval(dict_match.group())
                    chart_image = generate_chart("bar", chart_data, chart_data.get("title", "Analysis"))
        except:
            pass

        return {
            "reply": reply,
            "chart_image": chart_image,
            "mode": "analyze"
        }
    except Exception as e:
        return {"reply": f"Analysis error: {str(e)}"}

# ========== 6. FILE UPLOAD ==========
@app.post("/upload")
async def upload_file(file: UploadFile = File(...), session_id: str = Form("default")):
    try:
        content = await file.read()
        text = extract_text_from_file(content, file.filename)

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute(
            "INSERT INTO documents (session_id, filename, content, timestamp) VALUES (?, ?, ?, ?)",
            (session_id, file.filename, text, time.time())
        )
        conn.commit()
        conn.close()

        client = get_client()
        summary_prompt = f"""Summarize this document:

Filename: {file.filename}
Content (first 3000 chars): {text[:3000]}

Provide:
1. Brief summary
2. Key points
3. File type insights"""

        response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": summary_prompt}],
            temperature=0.5
        )

        reply = response.choices[0].message.content

        return {
            "reply": reply,
            "filename": file.filename,
            "size": len(content),
            "text_preview": text[:500]
        }
    except Exception as e:
        return {"reply": f"Upload error: {str(e)}"}

# ========== 7. IMAGE GENERATION ==========
@app.post("/generate-image")
async def generate_image_endpoint(request: ImageRequest):
    try:
        prompt = request.message
        images = []
        for i in range(3):
            img_url = generate_image(prompt, width=1024, height=768)
            images.append(img_url)

        return {
            "reply": f"Generated {len(images)} images for: '{prompt}'",
            "images": images,
            "mode": "image"
        }
    except Exception as e:
        return {"reply": f"Image generation error: {str(e)}"}

# ========== 7b. IMAGE SEARCH ==========
@app.post("/search-images")
async def search_images_endpoint(request: ImageSearchRequest):
    try:
        images = search_images(request.message, request.num_results)
        return {
            "reply": f"Found {len(images)} images for '{request.message}'",
            "images": images,
            "mode": "image_search"
        }
    except Exception as e:
        return {"reply": f"Image search error: {str(e)}"}

# ========== 8. WEB SEARCH ==========
@app.post("/search")
async def search_web(request: SearchRequest):
    try:
        results = web_search(request.message)
        return {"reply": results, "mode": "search"}
    except Exception as e:
        return {"reply": f"Search error: {str(e)}"}

# ========== 9. WEATHER ==========
@app.post("/weather")
async def weather_endpoint(request: WeatherRequest):
    try:
        city = request.message.strip()
        if not city:
            city = "London"
        result = get_weather(city)
        return {"reply": result, "mode": "weather"}
    except Exception as e:
        return {"reply": f"Weather error: {str(e)}"}

# ========== 10. NEWS ==========
@app.post("/news")
async def news_endpoint(request: NewsRequest):
    try:
        query = request.message.strip() or "technology"
        result = get_news(query)
        return {"reply": result, "mode": "news"}
    except Exception as e:
        return {"reply": f"News error: {str(e)}"}

# ========== 11. ACADEMIC SEARCH ==========
@app.post("/academic")
async def academic_endpoint(request: AcademicRequest):
    try:
        result = search_academic(request.message, request.max_results)
        return {"reply": result, "mode": "academic"}
    except Exception as e:
        return {"reply": f"Academic search error: {str(e)}"}

# ========== 12. FINANCIAL DATA ==========
@app.post("/financial")
async def financial_endpoint(request: FinancialRequest):
    try:
        symbol = request.symbol or request.message.strip().upper()
        symbol_match = re.search(r'[A-Z]{1,5}', request.message.upper())
        if symbol_match:
            symbol = symbol_match.group()
        result = get_financial_data(symbol, "quote")
        return {"reply": result, "mode": "financial", "symbol": symbol}
    except Exception as e:
        return {"reply": f"Financial data error: {str(e)}"}

# ========== 13. AGENT MODE (Enhanced) ==========
@app.post("/agent")
async def agent_mode(request: AgentRequest):
    try:
        client = get_client()

        tool_prompt = f"""You are an AI agent. Analyze this user request and decide which tools to use.

Available tools: search, code, weather, news, image, chart, translate, sentiment, academic, financial

User request: {request.message}

Respond with ONLY a JSON array of tool names to use, e.g.: ["search", "code"]
If no tools are needed, respond with: []"""

        tool_response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": tool_prompt}],
            temperature=0.3
        )

        tool_text = tool_response.choices[0].message.content.strip()

        try:
            tools_to_use = json.loads(tool_text)
        except:
            tools_to_use = []
            for tool in ["search", "code", "weather", "news", "image", "chart", "translate", "sentiment", "academic", "financial"]:
                if tool in tool_text.lower():
                    tools_to_use.append(tool)

        tool_results = {}

        if "search" in tools_to_use:
            tool_results["search"] = web_search(request.message)

        if "weather" in tools_to_use:
            cities = re.findall(r'in ([A-Za-z\s]+)', request.message)
            if cities:
                tool_results["weather"] = get_weather(cities[0])

        if "news" in tools_to_use:
            tool_results["news"] = get_news(request.message)

        if "academic" in tools_to_use:
            tool_results["academic"] = search_academic(request.message)

        if "financial" in tools_to_use:
            symbol_match = re.search(r'[A-Z]{1,5}', request.message.upper())
            if symbol_match:
                tool_results["financial"] = get_financial_data(symbol_match.group())

        if "code" in tools_to_use:
            code_prompt = f"Write Python code for: {request.message}"
            code_response = client.chat.completions.create(
                model="llama-3.1-8b-instant",
                messages=[{"role": "user", "content": code_prompt}],
                temperature=0.3
            )
            tool_results["code"] = code_response.choices[0].message.content

        synthesis_prompt = f"""Synthesize a helpful response based on the user request and tool results.

User request: {request.message}

Tools used: {tools_to_use}

Tool results:
{json.dumps(tool_results, indent=2)}

Provide a comprehensive, well-structured response."""

        final_response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": synthesis_prompt}],
            temperature=0.7,
            max_tokens=4096
        )

        reply = final_response.choices[0].message.content

        return {"reply": reply, "tools_used": tools_to_use, "mode": "agent"}
    except Exception as e:
        return {"reply": f"Agent error: {str(e)}"}

# ========== 14. TEXT TO SPEECH ==========
@app.post("/tts")
async def tts_endpoint(request: TTSRequest):
    try:
        audio = text_to_speech(request.message, request.lang)
        if audio:
            return {"reply": f"Text converted to speech: '{request.message[:100]}...'", "audio": audio, "mode": "tts"}
        else:
            return {"reply": "TTS not available. Install gTTS: pip install gtts", "mode": "tts"}
    except Exception as e:
        return {"reply": f"TTS error: {str(e)}"}

# ========== 15. CHART GENERATION ==========
@app.post("/chart")
async def chart_endpoint(request: ChartRequest):
    try:
        client = get_client()

        chart_prompt = f"""Extract chart data from this request and return a Python dictionary.

Request: {request.message}

Return ONLY a dictionary in this exact format:
{{"chart_type": "bar|line|pie|scatter|area", "labels": [...], "values": [...], "title": "...", "xlabel": "...", "ylabel": "..."}}

If the request is vague, create sample data that makes sense."""

        response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": chart_prompt}],
            temperature=0.3
        )

        chart_text = response.choices[0].message.content
        dict_match = re.search(r'\{.*\}', chart_text, re.DOTALL)
        if dict_match:
            chart_data = eval(dict_match.group())
            chart_type = chart_data.get("chart_type", "bar")
            chart_image = generate_chart(chart_type, chart_data, chart_data.get("title", "Chart"))

            return {
                "reply": f"Generated {chart_type} chart: {chart_data.get('title', 'Chart')}",
                "chart_image": chart_image,
                "mode": "chart"
            }
        else:
            return {"reply": "Could not parse chart data from request."}
    except Exception as e:
        return {"reply": f"Chart error: {str(e)}"}

# ========== 16. YOUTUBE VIDEO CREATOR ==========
@app.post("/create-video")
async def create_video(request: VideoRequest):
    try:
        client = get_client()

        video_prompt = f"""Create a YouTube video package for this topic:

Topic: {request.message}

Provide:
1. Catchy video title
2. SEO-optimized description (200 words)
3. 10 relevant tags
4. Video script outline (intro, 3 main points, outro with CTA)
5. Thumbnail text suggestion
6. Target audience

Format clearly with headers."""

        response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": video_prompt}],
            temperature=0.7,
            max_tokens=4096
        )

        reply = response.choices[0].message.content
        thumbnail_url = generate_image(f"YouTube thumbnail: {request.message}, bold text, eye-catching, professional, 16:9", width=1280, height=720)

        return {"reply": reply, "thumbnail_url": thumbnail_url, "mode": "video"}
    except Exception as e:
        return {"reply": f"Video creation error: {str(e)}"}

# ========== 17. TRANSLATION ==========
@app.post("/translate")
async def translate_endpoint(request: TranslateRequest):
    try:
        client = get_client()

        translate_prompt = f"""Translate the following text to {request.target_language}.

Text: {request.message}

Provide:
1. The translation
2. Pronunciation guide (if applicable)
3. Cultural notes or context
4. Alternative translations if ambiguous"""

        response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": translate_prompt}],
            temperature=0.3
        )

        reply = response.choices[0].message.content

        return {"reply": reply, "source_language": "auto-detected", "target_language": request.target_language, "mode": "translate"}
    except Exception as e:
        return {"reply": f"Translation error: {str(e)}"}

# ========== 18. SENTIMENT ANALYSIS ==========
@app.post("/sentiment")
async def sentiment_endpoint(request: SentimentRequest):
    try:
        client = get_client()

        sentiment_prompt = f"""Analyze the sentiment of this text comprehensively.

Text: {request.message}

Provide:
1. Overall sentiment (Positive/Negative/Neutral/Mixed) with confidence score
2. Emotion breakdown (joy, anger, sadness, fear, surprise, disgust) with percentages
3. Key phrases that indicate sentiment
4. Tone analysis (formal, casual, aggressive, supportive, etc.)
5. Actionable insights if this is customer feedback

Format as a structured report."""

        response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": sentiment_prompt}],
            temperature=0.3
        )

        reply = response.choices[0].message.content

        return {"reply": reply, "mode": "sentiment"}
    except Exception as e:
        return {"reply": f"Sentiment analysis error: {str(e)}"}

# ========== 19. CREATIVE WRITING ==========
@app.post("/creative")
async def creative_endpoint(request: CreativeRequest):
    try:
        client = get_client()

        genre_prompts = {
            "story": "Write an engaging, original short story with vivid characters, immersive setting, and compelling plot.",
            "poem": "Write an evocative, original poem with rich imagery, rhythm, and emotional depth.",
            "script": "Write a screenplay/script scene with dialogue, stage directions, and character development.",
            "essay": "Write a well-structured, persuasive essay with clear thesis, evidence, and conclusion.",
            "debate": "Write a structured debate argument with clear position, evidence, rebuttals, and conclusion."
        }

        creative_prompt = f"""{genre_prompts.get(request.genre, genre_prompts["story"])}

Topic/Prompt: {request.message}

Requirements:
- Be original and creative
- Use vivid, descriptive language
- Maintain consistent tone and style
- Engage the reader emotionally
- Show, don't just tell"""

        response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": creative_prompt}],
            temperature=0.9,
            max_tokens=4096
        )

        reply = response.choices[0].message.content

        return {"reply": reply, "genre": request.genre, "mode": "creative"}
    except Exception as e:
        return {"reply": f"Creative writing error: {str(e)}"}

# ========== 20. STUDY GUIDES ==========
@app.post("/study")
async def study_endpoint(request: StudyGuideRequest):
    try:
        client = get_client()

        study_prompt = f"""Create a comprehensive study guide for the following topic at {request.level} level.

Topic: {request.message}

Include:
1. Overview/Introduction
2. Key concepts with definitions
3. Important formulas/theories (if applicable)
4. Step-by-step explanations with analogies
5. Common misconceptions to avoid
6. Practice questions with answers
7. Summary/cheat sheet
8. Further resources/references

Adapt complexity to {request.level} level. Use analogies and examples."""

        response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": study_prompt}],
            temperature=0.5,
            max_tokens=4096
        )

        reply = response.choices[0].message.content

        return {"reply": reply, "level": request.level, "mode": "study"}
    except Exception as e:
        return {"reply": f"Study guide error: {str(e)}"}

# ========== 21. DEBATE & NEGOTIATION ==========
@app.post("/debate")
async def debate_endpoint(request: DebateRequest):
    try:
        client = get_client()

        debate_prompt = f"""You are a skilled debater. Analyze the following topic and present a well-reasoned argument.

Topic: {request.message}

Stance: {request.stance or "Analyze both sides (pro and con)"}

Structure:
1. Clear position statement
2. Main arguments with evidence
3. Counter-arguments and rebuttals
4. Consideration of multiple perspectives
5. Weighing of pros and cons
6. Conclusion with recommendation

Be persuasive, logical, and fair. Acknowledge limitations and uncertainties."""

        response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": debate_prompt}],
            temperature=0.7,
            max_tokens=4096
        )

        reply = response.choices[0].message.content

        return {"reply": reply, "stance": request.stance, "mode": "debate"}
    except Exception as e:
        return {"reply": f"Debate error: {str(e)}"}

# ========== 22. EMOTIONAL INTELLIGENCE ==========
@app.post("/emotional")
async def emotional_endpoint(request: ChatRequest):
    try:
        client = get_client()

        emotional_prompt = f"""You are an emotionally intelligent AI assistant. Analyze the following message and respond with empathy, understanding, and support.

User message: {request.message}

Instructions:
1. Identify the emotional tone (joy, sadness, anger, anxiety, excitement, etc.)
2. Acknowledge and validate the user's feelings
3. Respond with genuine empathy and compassion
4. Offer supportive, constructive guidance if appropriate
5. Be warm, non-judgmental, and personalized
6. Use an appropriate tone (calming, encouraging, celebratory, etc.)

Be human, caring, and authentic in your response."""

        response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": emotional_prompt}],
            temperature=0.8,
            max_tokens=4096
        )

        reply = response.choices[0].message.content

        # Save emotional context to memory
        save_memory(request.session_id, "last_emotion", request.message[:200])

        return {"reply": reply, "mode": "emotional"}
    except Exception as e:
        return {"reply": f"Emotional analysis error: {str(e)}"}

# ========== 23. RAG (DOCUMENT QA) ==========
@app.post("/rag/upload")
async def rag_upload(file: UploadFile = File(...), session_id: str = Form("default")):
    try:
        content = await file.read()
        text = extract_text_from_file(content, file.filename)

        if session_id not in rag_documents:
            rag_documents[session_id] = []

        rag_documents[session_id].append({
            "filename": file.filename,
            "content": text,
            "timestamp": time.time()
        })

        create_rag_index(session_id, rag_documents[session_id])

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute(
            "INSERT INTO documents (session_id, filename, content, timestamp) VALUES (?, ?, ?, ?)",
            (session_id, file.filename, text, time.time())
        )
        conn.commit()
        conn.close()

        return {
            "reply": f"Document '{file.filename}' uploaded and indexed. You can now ask questions about it in RAG mode.",
            "filename": file.filename,
            "chunks": len(text) // 500
        }
    except Exception as e:
        return {"reply": f"RAG upload error: {str(e)}"}

@app.post("/rag/query")
async def rag_query(request: RAGQueryRequest):
    try:
        session_id = request.session_id or "default"

        rag_results = search_rag(session_id, request.message, top_k=5)
        docs = rag_documents.get(session_id, [])

        if not docs:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            c.execute(
                "SELECT filename, content FROM documents WHERE session_id = ? ORDER BY timestamp DESC",
                (session_id,)
            )
            rows = c.fetchall()
            conn.close()
            docs = [{"filename": r[0], "content": r[1]} for r in rows]

        if not docs:
            return {"reply": "No documents found. Please upload documents first in RAG mode."}

        if rag_results:
            context_parts = [f"From {fname}:\n{chunk}" for fname, chunk in rag_results]
            all_content = "\n\n".join(context_parts)
        else:
            all_content = "\n\n".join([f"--- {d['filename']} ---\n{d['content'][:5000]}" for d in docs])

        client = get_client()

        rag_prompt = f"""Answer the following question based ONLY on the provided documents.

Documents:
{all_content[:15000]}

Question: {request.message}

Instructions:
- Answer based only on the documents provided
- If the answer is not in the documents, say so clearly
- Cite which document(s) you used
- Be concise but thorough"""

        response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": rag_prompt}],
            temperature=0.3,
            max_tokens=4096
        )

        reply = response.choices[0].message.content

        return {
            "reply": reply,
            "documents_used": [d["filename"] for d in docs],
            "mode": "rag"
        }
    except Exception as e:
        return {"reply": f"RAG query error: {str(e)}"}

# ========== 24. CLEAR HISTORY ==========
@app.post("/clear")
async def clear_history(request: ClearRequest):
    session_id = request.session_id or "default"

    if session_id in conversations:
        conversations[session_id] = [
            {"role": "system", "content": SYSTEM_PROMPTS["chat"]}
        ]

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM conversations WHERE session_id = ?", (session_id,))
    conn.commit()
    conn.close()

    return {"status": "cleared", "session_id": session_id}

# ========== 25. GET HISTORY ==========
@app.get("/history/{session_id}")
async def get_history(session_id: str):
    history = get_conversation_history(session_id, limit=50)
    return {"history": history, "session_id": session_id}

# ========== 26. HEALTH CHECK ==========
@app.get("/health")
async def health_check():
    return {
        "status": "healthy",
        "version": "3.0.0",
        "features_active": {
            "groq": bool(GROQ_API_KEY),
            "web_search": bool(TAVILY_API_KEY),
            "weather": bool(OPENWEATHER_API_KEY),
            "news": bool(NEWSDATA_API_KEY),
            "tts": GTTS_AVAILABLE,
            "pdf": PYPDF_AVAILABLE,
            "excel": OPENPYXL_AVAILABLE,
            "faiss": FAISS_AVAILABLE,
            "sentence_transformers": SENTENCE_TRANSFORMERS_AVAILABLE
        }
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=int(os.getenv("PORT", 10000)))